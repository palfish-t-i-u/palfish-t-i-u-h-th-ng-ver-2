"""Fix 12/8: xuất HĐ B4 sai "Tên sản phẩm" (hiện tên khách) + "Địa chỉ" trống.

Root cause: course["name"] (tên gói học, set lúc tạo AR) và tên khách hàng để
xuất HĐ đều dùng chung key "name" trong _build_invoice_course_patch — patch
tên khách ghi đè mất tên gói. Fix: tên khách dùng key riêng
"invoice_customer_name", "name" chỉ còn giữ đúng nghĩa tên gói học xuyên suốt.
Đồng thời bổ sung "diaChi" vào _course_to_tax_order() (trước đây không có,
cột 5 file khach_hang.xlsx luôn trống).

Xem docs/plans/PLAN_FIX_INVOICE_EXPORT_2026-08-12.md.
"""
from __future__ import annotations

import activation_routes as ar

FULL_PR = {
    "name": "Nguyễn Văn A",
    "phone": "0900000000",
    "province": "Thành phố Hà Nội",
    "ward": "Phường Hoàng Mai",
    "address": "119 Phúc Xá",
}


def _course(**over):
    base = {
        "code": "CC-0078-001",
        "name": "2/W- NEW 24 buoi",
        "amount": 4_550_000,
        "order_id": "752820050659",
        "invoiced": False,
    }
    base.update(over)
    return base


def test_issue_invoice_patch_does_not_clobber_package_name():
    """_build_invoice_course_patch ghi tên khách vào invoice_customer_name,
    KHÔNG được đụng vào "name" (tên gói học) của course."""
    course = _course()
    body = ar.IssueCourseInvoiceBody(
        name="Trần Thị B",
        phone="0911111111",
        province="Thành phố Hà Nội",
        ward="Phường Hoàng Mai",
        address="119 Phúc Xá",
    )
    patch = ar._build_invoice_course_patch(course, None, body)

    assert patch["invoice_customer_name"] == "Trần Thị B"
    assert "name" not in patch  # không được ghi đè key "name"


def test_course_to_tax_order_keeps_package_name_after_invoice_issued():
    """Course đã qua bước xuất HĐ (đã áp patch tên khách) — B4 export vẫn phải
    ra đúng tên gói học, không phải tên khách."""
    course = _course()
    body = ar.IssueCourseInvoiceBody(
        name="Trần Thị B",
        phone="0911111111",
        province="Thành phố Hà Nội",
        ward="Phường Hoàng Mai",
        address="119 Phúc Xá",
    )
    patch = ar._build_invoice_course_patch(course, None, body)
    course.update(patch)  # mô phỏng đúng những gì _issue_course_invoice_atomic làm với course dict

    order = ar._course_to_tax_order(course, {}, {"customer_name": ""}, None, "M260812001", "PF000001")

    assert order["taxProductName"] == "2/W- NEW 24 buoi"
    assert order["goiHoc"] == "2/W- NEW 24 buoi"
    assert order["tenKhach"] == "Trần Thị B"


def test_course_display_name_reads_invoice_customer_name_not_package_name():
    course = _course(invoice_customer_name="Trần Thị B")
    assert ar._course_display_name(course, {"customer_name": ""}, None) == "Trần Thị B"


def test_course_display_name_falls_back_when_no_invoice_customer_name():
    course = _course()  # chưa qua bước xuất HĐ — chưa có invoice_customer_name
    assert ar._course_display_name(course, {"customer_name": "Lê Văn C"}, None) == "Lê Văn C"


def test_dia_chi_uses_course_then_pr_fallback():
    course = _course(province="Thành phố Hà Nội", ward="Phường Hoàng Mai", address="119 Phúc Xá")
    assert ar._course_full_address(course, None) == "119 Phúc Xá, Phường Hoàng Mai, Thành phố Hà Nội"

    course_no_addr = _course()
    assert ar._course_full_address(course_no_addr, FULL_PR) == "119 Phúc Xá, Phường Hoàng Mai, Thành phố Hà Nội"


def test_dia_chi_empty_when_no_address_anywhere():
    assert ar._course_full_address(_course(), None) == ""


def test_course_to_tax_order_includes_dia_chi():
    course = _course()
    order = ar._course_to_tax_order(course, {}, {"customer_name": ""}, FULL_PR, "M260812001", "PF000001")
    assert order["diaChi"] == "119 Phúc Xá, Phường Hoàng Mai, Thành phố Hà Nội"


def test_build_excel_customers_writes_dia_chi_column_5():
    import io

    import openpyxl

    import invoice_routes as inv

    orders = [{
        "sdt": "0900000000",
        "tenKhach": "Trần Thị B",
        "diaChi": "119 Phúc Xá, Phường Hoàng Mai, Thành phố Hà Nội",
    }]
    data = inv._build_excel_customers(orders)
    wb = openpyxl.load_workbook(io.BytesIO(data))
    ws = wb.active

    assert ws.cell(row=1, column=5).value == "Địa chỉ"
    assert ws.cell(row=2, column=3).value == "Trần Thị B"
    assert ws.cell(row=2, column=5).value == "119 Phúc Xá, Phường Hoàng Mai, Thành phố Hà Nội"


def test_build_excel_customers_dia_chi_blank_when_missing():
    import io

    import openpyxl

    import invoice_routes as inv

    orders = [{"sdt": "0900000000", "tenKhach": "Trần Thị B"}]
    data = inv._build_excel_customers(orders)
    wb = openpyxl.load_workbook(io.BytesIO(data))
    ws = wb.active

    assert ws.cell(row=2, column=5).value is None


# ── CCCD/MST/email trên file 02 (thuế NĐ 70/2025, chuẩn Sương Mai 26/8) ──────

def test_course_to_tax_order_so_cccd_ca_nhan_tu_pr_tax_id():
    pr = {**FULL_PR, "tax_id": "001204012345", "email": "a@example.com"}
    order = ar._course_to_tax_order(_course(), {}, {"customer_name": ""}, pr, "M1", "PF1")
    assert order["soCccd"] == "001204012345"
    assert order["maSoThue"] == ""
    assert order["tenDonVi"] == ""
    assert order["email"] == "a@example.com"


def test_course_to_tax_order_tax_code_per_course_uu_tien_hon_pr():
    pr = {**FULL_PR, "tax_id": "001204012345"}
    order = ar._course_to_tax_order(
        _course(tax_code="079304005678"), {}, {"customer_name": ""}, pr, "M1", "PF1"
    )
    assert order["soCccd"] == "079304005678"


def test_course_to_tax_order_doanh_nghiep_mst_va_ten_don_vi():
    pr = {
        **FULL_PR,
        "customer_type": "business",
        "tax_id": "0123456789",
        "company_name": "Công ty TNHH ABC",
    }
    order = ar._course_to_tax_order(_course(), {}, {"customer_name": ""}, pr, "M1", "PF1")
    assert order["maSoThue"] == "0123456789"
    assert order["tenDonVi"] == "Công ty TNHH ABC"
    assert order["soCccd"] == ""


def test_course_display_name_uu_tien_ho_ten_phap_ly_tren_pr():
    """PR.invoice_customer_name (tên pháp lý sale khai) thắng pr["name"] ("Chị Hằng")."""
    pr = {**FULL_PR, "name": "Chị Hằng", "invoice_customer_name": "Nguyễn Thị Hằng"}
    assert ar._course_display_name(_course(), {"customer_name": ""}, pr) == "Nguyễn Thị Hằng"


def test_build_excel_customers_writes_cccd_mst_email_columns():
    import io

    import openpyxl

    import invoice_routes as inv

    orders = [{
        "sdt": "0900000000",
        "tenKhach": "Nguyễn Thị Hằng",
        "diaChi": "Phường Hoàng Mai, Thành phố Hà Nội",
        "soCccd": "001204012345",
        "maSoThue": "",
        "tenDonVi": "",
        "email": "hang@example.com",
    }]
    data = inv._build_excel_customers(orders)
    wb = openpyxl.load_workbook(io.BytesIO(data))
    ws = wb.active

    assert ws.cell(row=1, column=8).value == "Số CCCD"
    assert ws.cell(row=1, column=9).value == "Mã số thuế"
    assert ws.cell(row=2, column=8).value == "001204012345"
    assert ws.cell(row=2, column=10).value == "hang@example.com"


def test_build_invoice_course_patch_fallback_tax_va_email_tu_pr():
    pr = {
        **FULL_PR,
        "wants_invoice": True,
        "invoice_customer_name": "Nguyễn Văn A",
        "tax_id": "001204012345",
        "email": "a@example.com",
    }
    patch = ar._build_invoice_course_patch(_course(), pr, None)
    assert patch["tax_code"] == "001204012345"
    assert patch["email"] == "a@example.com"
    assert patch["invoice_customer_name"] == "Nguyễn Văn A"


def test_build_invoice_course_patch_taker_thieu_cccd_bi_chan():
    import pytest
    from fastapi import HTTPException

    pr = {**FULL_PR, "wants_invoice": True, "invoice_customer_name": "Nguyễn Văn A", "email": "a@x.com"}
    with pytest.raises(HTTPException) as exc:
        ar._build_invoice_course_patch(_course(), pr, None)
    assert "CCCD" in str(exc.value.detail)


def test_build_invoice_course_patch_taker_ho_ten_khong_tinh_pr_name():
    """pr["name"] fallback cho hiển thị nhưng KHÔNG thỏa gate họ tên đầy đủ."""
    import pytest
    from fastapi import HTTPException

    pr = {**FULL_PR, "wants_invoice": True, "tax_id": "001204012345", "email": "a@x.com"}
    with pytest.raises(HTTPException) as exc:
        ar._build_invoice_course_patch(_course(), pr, None)
    assert "họ tên" in str(exc.value.detail)


def test_build_invoice_course_patch_non_taker_khong_gate():
    """Khách không lấy HĐ: thiếu hết CCCD/email/địa chỉ vẫn xuất được (chỉ cần tên+SĐT)."""
    pr = {"name": "Chị Hằng", "phone": "0900000000"}
    patch = ar._build_invoice_course_patch(_course(), pr, None)
    assert patch["invoice_customer_name"] == "Chị Hằng"


def test_build_invoice_course_patch_grandfather_don_cu_khong_chan():
    """Đơn taker tạo trước live CCCD (27/8 VN) — phát hành HĐ không bị 400 dù thiếu CCCD/email."""
    pr = {
        **FULL_PR,
        "wants_invoice": True,
        "created_at": "2026-08-22T10:00:00+00:00",
    }
    patch = ar._build_invoice_course_patch(_course(), pr, None)
    assert patch["invoice_customer_name"] == "Nguyễn Văn A"  # fallback hiển thị pr["name"]


def test_build_invoice_course_patch_taker_so_nha_khong_bat_buoc():
    pr = {
        "name": "Chị Hằng",
        "phone": "0900000000",
        "wants_invoice": True,
        "invoice_customer_name": "Nguyễn Thị Hằng",
        "tax_id": "001204012345",
        "email": "a@x.com",
        "province": "Thành phố Hà Nội",
        "ward": "Phường Hoàng Mai",
        "address": "",
    }
    patch = ar._build_invoice_course_patch(_course(), pr, None)
    assert patch["invoice_customer_name"] == "Nguyễn Thị Hằng"
