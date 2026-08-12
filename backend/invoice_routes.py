"""Invoice routes — Module 3 (Xác nhận CRM) + Module 4 (Xuất hóa đơn thuế).

Luồng:
  tien_ve=True → trang_thai_thu_tuc='CHO_XAC_NHAN'  (set bởi patch_order/trigger)
       ↓  [M3 – Ops xác nhận]
  trang_thai_thu_tuc='CHO_XUAT_HD'
       ↓  [M4 – Ops bấm Tải hóa đơn]
  trang_thai_thu_tuc='DA_XUAT_HD'  +  mã M.../PF... được chốt cứng
"""

from __future__ import annotations

import io
import os
import tempfile
import zipfile
from datetime import date, datetime, timezone
from typing import Any

from fastapi import Header, HTTPException
from fastapi.responses import FileResponse
from pydantic import BaseModel, Field
from starlette.background import BackgroundTask

from rbac import can_confirm_payment, resolve_actor
from revenue_routes import sync_ledger_from_m3_order

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
TRANG_THAI_CHO_XAC_NHAN = "CHO_XAC_NHAN"
TRANG_THAI_CHO_XUAT_HD = "CHO_XUAT_HD"
TRANG_THAI_DA_XUAT_HD = "DA_XUAT_HD"


# ---------------------------------------------------------------------------
# Pydantic request bodies
# ---------------------------------------------------------------------------
class M3ApproveBody(BaseModel):
    id: str
    taxProductName: str
    crmOrderId: str = ""


class M3SaveBody(BaseModel):
    id: str
    crmOrderId: str = ""


class M4CancelBody(BaseModel):
    id: str


class ExportBatchBody(BaseModel):
    order_ids: list[str] = Field(default_factory=list)


# ---------------------------------------------------------------------------
# RBAC helper
# ---------------------------------------------------------------------------
def _require_ops(actor) -> None:
    if not can_confirm_payment(actor):
        raise HTTPException(403, "Chỉ Ops/System được thực hiện thao tác hóa đơn")


# ---------------------------------------------------------------------------
# Serializer
# ---------------------------------------------------------------------------
def _row_to_invoice_order(row: dict[str, Any], kh: dict[str, Any] | None = None) -> dict[str, Any]:
    k = kh or {}
    sdt_raw = k.get("so_dien_thoai") or row.get("sdt") or ""
    ma_vung = (k.get("ma_vung") or "+84").strip()
    sdt = f"{ma_vung} {sdt_raw}" if sdt_raw and not str(sdt_raw).startswith("+") else sdt_raw
    return {
        "id": row["id"],
        "maDonHang": row.get("ma_don_hang") or "",
        "tenKhach": k.get("ho_ten") or row.get("ten_khach") or "",
        "sdt": sdt,
        "uid": k.get("crm_uid") or "",
        "goiHoc": row.get("goi_hoc") or "",
        "tongTien": int(row.get("so_tien_can_thu") or 0),
        "nguon": row.get("nguon_doanh_thu") or "",
        "tienVe": bool(row.get("tien_ve")),
        "donCRM": bool(row.get("don_crm")),
        "trangThaiThuTuc": row.get("trang_thai_thu_tuc") or "",
        "taxProductName": row.get("tax_product_name") or "",
        "taxInvoiceCode": row.get("tax_invoice_code") or "",
        "taxProductCode": row.get("tax_product_code") or "",
        "m3ApprovedAt": row.get("m3_approved_at") or "",
        "crmOrderId": row.get("crm_order_id") or "",
        "createdBy": row.get("created_by") or "",
        "createdAt": row.get("created_at") or "",
    }


# ---------------------------------------------------------------------------
# Sequence allocation — batch-safe for single-operator flow
# ---------------------------------------------------------------------------
def _alloc_sequences(sb, n: int, date_key: str) -> tuple[int, int]:
    """Allocate n invoice sequence numbers (daily-reset) + n product codes (global).

    Returns (inv_start, prod_start).  Caller assigns:
      invoice code i  →  f"M{date_key}{inv_start+i+1:03d}"
      product code i  →  f"PF{prod_start+i+1:06d}"
    """
    from rpc_helpers import rpc_allocate_tax_sequences

    return rpc_allocate_tax_sequences(sb, n, date_key)


# ---------------------------------------------------------------------------
# Excel builders — 3 files in memory (openpyxl)
# ---------------------------------------------------------------------------
def _openpyxl_imports():
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter

        return openpyxl, Font, PatternFill, Alignment, Border, Side, get_column_letter
    except ImportError as exc:
        raise HTTPException(500, f"Thiếu thư viện openpyxl — pip install openpyxl: {exc}") from exc


def _fmt_sdt(raw: str) -> str:
    """Chuẩn hoá SĐT sang dạng 84-XXXXXXXXX."""
    digits = "".join(c for c in (raw or "") if c.isdigit())
    if digits.startswith("84") and len(digits) == 11:
        return f"84-{digits[2:]}"
    if digits.startswith("0") and len(digits) == 10:
        return f"84-{digits[1:]}"
    if len(digits) == 9:
        return f"84-{digits}"
    return raw or ""


def _apply_cell_style(cell, font, alignment, border):
    cell.font = font
    cell.alignment = alignment
    cell.border = border


def _build_excel_orders(orders: list[dict[str, Any]]) -> bytes:
    """File 1: don_hang.xlsx — 2-dòng header với merge cells theo mẫu kế toán."""
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise HTTPException(500, f"Thiếu thư viện openpyxl: {exc}") from exc

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Đơn hàng"

    thin = Side(style="thin", color="000000")
    bord = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left   = Alignment(horizontal="left",   vertical="center", wrap_text=False)

    # ── Dòng 1: nhóm gộp ô ──────────────────────────────────────────────────
    groups = [
        ("A1:G1",  "Thông tin chung",    "4472C4"),
        ("H1:O1",  "Thông tin sản phẩm", "70AD47"),
        ("P1:Q1",  "Thông tin lô hàng",  "ED7D31"),
    ]
    for rng, label, color in groups:
        ws.merge_cells(rng)
        first_col = rng.split(":")[0]
        cell = ws[first_col]
        cell.value = label
        cell.font      = Font(bold=True, color="FFFFFF", size=11)
        cell.fill      = PatternFill("solid", fgColor=color)
        cell.alignment = center
        cell.border    = bord
    # border cho các ô bị gộp (openpyxl không tự áp border vào ô con)
    for rng, _, color in groups:
        start_col_letter, end_col_letter = rng.split(":")[0][0], rng.split(":")[1][0]
        start_ci = ord(start_col_letter) - ord("A") + 1
        end_ci   = ord(end_col_letter)   - ord("A") + 1
        for ci in range(start_ci, end_ci + 1):
            c = ws.cell(row=1, column=ci)
            c.border = bord
    ws.row_dimensions[1].height = 22

    # ── Dòng 2: tên cột ──────────────────────────────────────────────────────
    # Chú ý: "Thuế suất " có dấu cách cuối — giữ nguyên theo spec
    headers_r2 = [
        "Mã đơn hàng*",          # A
        "Mã khách hàng",          # B
        "Tên khách hàng*",        # C
        "Tên người mua",          # D
        "Ngày hóa đơn",           # E
        "Hình thức thanh toán",   # F
        "Ghi chú",                # G
        "Mã sản phẩm",            # H
        "Tên sản phẩm*",          # I
        "Đơn vị tính",            # J
        "Số lượng*",              # K
        "Đơn giá*",               # L
        "Giảm giá sản phẩm",      # M
        "Thuế suất ",             # N  ← dấu cách cuối theo spec
        "Tiền sau thuế",          # O
        "Mã lô*",                 # P
        "Lượng hàng*",            # Q
    ]
    col_widths = [18, 18, 22, 22, 16, 20, 14, 16, 36, 12, 10, 16, 18, 12, 16, 12, 12]
    h2_fill = PatternFill("solid", fgColor="D9D9D9")
    h2_font = Font(bold=True, size=10)
    for ci, (h, w) in enumerate(zip(headers_r2, col_widths), 1):
        cell = ws.cell(row=2, column=ci, value=h)
        cell.fill      = h2_fill
        cell.font      = h2_font
        cell.alignment = center
        cell.border    = bord
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[2].height = 28

    # ── Data từ dòng 3 ────────────────────────────────────────────────────────
    data_font = Font(size=10)
    for ri, o in enumerate(orders, 3):
        inv_date = ""
        if o.get("m3ApprovedAt"):
            try:
                dt = datetime.fromisoformat(str(o["m3ApprovedAt"]).replace("Z", "+00:00"))
                inv_date = dt.strftime("%Y-%m-%d")
            except Exception:
                inv_date = str(o["m3ApprovedAt"])[:10]

        row_vals = [
            o.get("taxInvoiceCode", ""),   # A  Mã đơn hàng*
            _fmt_sdt(o.get("sdt", "")),    # B  Mã khách hàng
            "",                            # C  Tên khách hàng* (để trống)
            "",                            # D  Tên người mua (để trống)
            inv_date,                      # E  Ngày hóa đơn
            "TM/CK",                       # F  Hình thức thanh toán
            "",                            # G  Ghi chú
            o.get("taxProductCode", ""),   # H  Mã sản phẩm
            o.get("taxProductName", "") or o.get("goiHoc", ""),  # I  Tên sản phẩm*
            "Khóa",                        # J  Đơn vị tính
            1,                             # K  Số lượng*
            int(o.get("tongTien", 0)),     # L  Đơn giá*
            "",                            # M  Giảm giá sản phẩm
            "KCT",                         # N  Thuế suất
            "",                            # O  Tiền sau thuế
            "",                            # P  Mã lô*
            "",                            # Q  Lượng hàng*
        ]
        for ci, val in enumerate(row_vals, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font      = data_font
            cell.border    = bord
            cell.alignment = center if ci in (10, 11) else left
            if ci == 12:  # Đơn giá — format số
                cell.number_format = "#,##0"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_excel_customers(orders: list[dict[str, Any]]) -> bytes:
    """File 2: khach_hang.xlsx — flat header, deduplicate theo Mã KH (84-SĐT)."""
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise HTTPException(500, f"Thiếu thư viện openpyxl: {exc}") from exc

    # Deduplicate theo 84-SĐT (Mã KH)
    seen: set[str] = set()
    unique_rows: list[dict[str, Any]] = []
    for o in orders:
        ma_kh = _fmt_sdt(o.get("sdt", ""))
        if ma_kh not in seen:
            seen.add(ma_kh)
            unique_rows.append({**o, "_maKH": ma_kh})

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Khách hàng"

    thin  = Side(style="thin", color="000000")
    bord  = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    left   = Alignment(horizontal="left",   vertical="center")

    headers = [
        "Mã KH/NCC",          # 1
        "Tên đơn vị (*)",     # 2
        "Tên người mua",      # 3
        "Loại (*)",           # 4
        "Địa chỉ",            # 5
        "Số điện thoại",      # 6
        "Ngày sinh",          # 7
        "Số CCCD",            # 8
        "Mã số thuế",         # 9
        "Email",              # 10
        "Số tài khoản",       # 11
        "Tên ngân hàng",      # 12
        "Giới tính",          # 13
        "Mô tả",              # 14
    ]
    col_widths = [20, 30, 24, 14, 30, 18, 12, 16, 16, 24, 18, 20, 10, 20]

    h_fill = PatternFill("solid", fgColor="70AD47")
    h_font = Font(bold=True, color="FFFFFF", size=10)
    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill      = h_fill
        cell.font      = h_font
        cell.alignment = center
        cell.border    = bord
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 26

    data_font = Font(size=10)
    for ri, row in enumerate(unique_rows, 2):
        # 14 cột; điền 1, 3, 4, 5 — còn lại rỗng
        row_vals = [
            row["_maKH"],           # 1  Mã KH/NCC = 84-SĐT
            "",                     # 2  Tên đơn vị
            row.get("tenKhach", ""),# 3  Tên người mua = họ tên khách
            "Khách hàng",           # 4  Loại
            row.get("diaChi", ""),  # 5  Địa chỉ
            "", "", "", "", "", "", "", "", "",  # 6–14
        ]
        for ci, val in enumerate(row_vals, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font      = data_font
            cell.border    = bord
            cell.alignment = left

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_excel_products(orders: list[dict[str, Any]]) -> bytes:
    """File 3: san_pham.xlsx — flat header, 1 dòng / đơn, không lọc trùng."""
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise HTTPException(500, f"Thiếu thư viện openpyxl: {exc}") from exc

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sản phẩm"

    thin  = Side(style="thin", color="000000")
    bord  = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    left   = Alignment(horizontal="left",   vertical="center")

    headers = [
        "Mã sản phẩm (*)",              # 1
        "Tên sản phẩm (*)",             # 2
        "Barcode",                       # 3
        "Nhóm sản phẩm",                # 4
        "Giá nhập",                      # 5
        "Giá bán (*)",                   # 6
        "Đơn vị tính",                   # 7
        "Theo dõi hàng tồn kho",         # 8
        "Số lượng tồn kho",              # 9
        "Hạn mức tồn",                   # 10
        "Áp dụng thuế",                  # 11
        "Ảnh",                           # 12
        "Mô tả",                         # 13
        "Thuế giảm trừ",                 # 14
        "Nhóm ngành nghề tính thuế",     # 15
    ]
    col_widths = [18, 36, 14, 18, 14, 16, 14, 22, 18, 12, 14, 10, 20, 16, 28]

    h_fill = PatternFill("solid", fgColor="ED7D31")
    h_font = Font(bold=True, color="FFFFFF", size=10)
    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill      = h_fill
        cell.font      = h_font
        cell.alignment = center
        cell.border    = bord
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 26

    data_font = Font(size=10)
    for ri, o in enumerate(orders, 2):
        row_vals = [
            o.get("taxProductCode", ""),                            # 1  Mã sản phẩm (*)
            o.get("taxProductName", "") or o.get("goiHoc", ""),    # 2  Tên sản phẩm (*)
            "",                                                      # 3  Barcode
            "",                                                      # 4  Nhóm sản phẩm
            "",                                                      # 5  Giá nhập
            int(o.get("tongTien", 0)),                              # 6  Giá bán (*)
            "Khóa",                                                  # 7  Đơn vị tính
            "Không",                                                 # 8  Theo dõi hàng tồn kho
            "", "", "", "", "", "", "",                              # 9–15
        ]
        for ci, val in enumerate(row_vals, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font      = data_font
            cell.border    = bord
            cell.alignment = left
            if ci == 6:
                cell.number_format = "#,##0"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Route registration — same closure pattern as admin_routes.py
# ---------------------------------------------------------------------------
def register_invoice_routes(app, get_supabase) -> None:
    """Attach /invoice/* routes to the FastAPI app."""

    def _sb():
        sb = get_supabase()
        if not sb:
            raise HTTPException(503, "Supabase chưa cấu hình")
        return sb

    # ------------------------------------------------------------------
    # GET /invoice/m3-pending
    # ------------------------------------------------------------------
    @app.get("/invoice/m3-pending")
    def m3_pending(authorization: str | None = Header(None)):
        """Tất cả đơn đã có tiền về — bao gồm cả đã xác nhận / đã xuất để hiển thị lịch sử."""
        sb = _sb()
        actor = resolve_actor(sb, authorization)
        _require_ops(actor)
        try:
            res = (
                sb.table("don_hang")
                .select("*, khach_hang(*)")
                .eq("tien_ve", True)
                .order("created_at", desc=True)
                .execute()
            )
            out = []
            for row in res.data or []:
                kh = row.pop("khach_hang", None) if isinstance(row, dict) else None
                out.append(_row_to_invoice_order(row, kh))
            return {"orders": out, "count": len(out)}
        except HTTPException:
            raise
        except Exception as exc:
            raise HTTPException(500, f"Lỗi lấy danh sách M3: {exc}") from exc

    # ------------------------------------------------------------------
    # POST /invoice/m3-save  — lưu crmOrderId không đổi trạng thái
    # ------------------------------------------------------------------
    @app.post("/invoice/m3-save")
    def m3_save(body: M3SaveBody, authorization: str | None = Header(None)):
        """Lưu crmOrderId cho đơn mà không thay đổi trang_thai_thu_tuc."""
        sb = _sb()
        actor = resolve_actor(sb, authorization)
        _require_ops(actor)
        try:
            res = (
                sb.table("don_hang")
                .update({
                    "crm_order_id": body.crmOrderId.strip() or None,
                    "don_crm": True,
                })
                .eq("id", body.id)
                .execute()
            )
            if not res.data:
                raise HTTPException(404, "Không tìm thấy đơn hàng")
            row = res.data[0]
            kh_res = (
                sb.table("khach_hang")
                .select("*")
                .eq("id", row.get("khach_hang_id"))
                .limit(1)
                .execute()
            )
            return _row_to_invoice_order(row, kh_res.data[0] if kh_res.data else None)
        except HTTPException:
            raise
        except Exception as exc:
            raise HTTPException(500, f"Lỗi lưu CRM ID: {exc}") from exc

    # ------------------------------------------------------------------
    # POST /invoice/m3-approve
    # ------------------------------------------------------------------
    @app.post("/invoice/m3-approve")
    def m3_approve(body: M3ApproveBody, authorization: str | None = Header(None)):
        """Ops xác nhận: điền taxProductName + crmOrderId, chuyển sang CHO_XUAT_HD."""
        sb = _sb()
        actor = resolve_actor(sb, authorization)
        _require_ops(actor)

        if not body.taxProductName.strip():
            raise HTTPException(400, "taxProductName không được để trống")

        try:
            patch: dict[str, Any] = {
                "tax_product_name": body.taxProductName.strip(),
                "crm_order_id": body.crmOrderId.strip() or None,
                "m3_approved_at": datetime.now(timezone.utc).isoformat(),
                "trang_thai_thu_tuc": TRANG_THAI_CHO_XUAT_HD,
            }
            res = (
                sb.table("don_hang")
                .update(patch)
                .eq("id", body.id)
                .eq("trang_thai_thu_tuc", TRANG_THAI_CHO_XAC_NHAN)
                .execute()
            )
            if not res.data:
                raise HTTPException(
                    404, "Đơn không tìm thấy hoặc không còn ở trạng thái CHO_XAC_NHAN"
                )
            row = res.data[0]
            kh_res = (
                sb.table("khach_hang")
                .select("*")
                .eq("id", row.get("khach_hang_id"))
                .limit(1)
                .execute()
            )
            sync_ledger_from_m3_order(sb, body.id, actor.email)
            return _row_to_invoice_order(row, kh_res.data[0] if kh_res.data else None)
        except HTTPException:
            raise
        except Exception as exc:
            raise HTTPException(500, f"Lỗi M3 approve: {exc}") from exc

    # ------------------------------------------------------------------
    # POST /invoice/m3-approve-bulk
    # ------------------------------------------------------------------
    @app.post("/invoice/m3-approve-bulk")
    def m3_approve_bulk(authorization: str | None = Header(None)):
        """Bulk-approve tất cả đơn CHO_XAC_NHAN → CHO_XUAT_HD.

        Dùng goi_hoc làm tax_product_name (fallback: ma_don_hang).
        Trả về số đơn được duyệt.
        """
        sb = _sb()
        actor = resolve_actor(sb, authorization)
        _require_ops(actor)

        try:
            res = (
                sb.table("don_hang")
                .select("id, goi_hoc, ma_don_hang, crm_order_id")
                .eq("trang_thai_thu_tuc", TRANG_THAI_CHO_XAC_NHAN)
                .eq("tien_ve", True)
                .execute()
            )
        except Exception as exc:
            raise HTTPException(500, f"Lỗi đọc danh sách CHO_XAC_NHAN: {exc}") from exc

        rows = res.data or []
        if not rows:
            return {"approved": 0, "ids": []}

        now_iso = datetime.now(timezone.utc).isoformat()
        approved_ids: list[str] = []
        for row in rows:
            tax_name = (row.get("goi_hoc") or row.get("ma_don_hang") or "").strip()
            if not tax_name:
                continue
            try:
                upd = (
                    sb.table("don_hang")
                    .update({
                        "tax_product_name": tax_name,
                        "m3_approved_at": now_iso,
                        "trang_thai_thu_tuc": TRANG_THAI_CHO_XUAT_HD,
                    })
                    .eq("id", row["id"])
                    .eq("trang_thai_thu_tuc", TRANG_THAI_CHO_XAC_NHAN)
                    .execute()
                )
                if upd.data:
                    approved_ids.append(row["id"])
                    sync_ledger_from_m3_order(sb, row["id"], actor.email)
            except Exception:
                pass  # skip đơn lỗi, tiếp tục các đơn còn lại

        return {"approved": len(approved_ids), "ids": approved_ids}

    # ------------------------------------------------------------------
    # GET /invoice/m4-queue
    # ------------------------------------------------------------------
    @app.get("/invoice/m4-queue")
    def m4_queue(authorization: str | None = Header(None)):
        """Queue các đơn đã qua M3, đang chờ xuất hóa đơn thuế."""
        sb = _sb()
        actor = resolve_actor(sb, authorization)
        _require_ops(actor)
        try:
            res = (
                sb.table("don_hang")
                .select("*, khach_hang(*)")
                .eq("trang_thai_thu_tuc", TRANG_THAI_CHO_XUAT_HD)
                .order("m3_approved_at", desc=False)
                .execute()
            )
            out = []
            for row in res.data or []:
                kh = row.pop("khach_hang", None) if isinstance(row, dict) else None
                out.append(_row_to_invoice_order(row, kh))
            return {"orders": out, "count": len(out)}
        except HTTPException:
            raise
        except Exception as exc:
            raise HTTPException(500, f"Lỗi lấy queue M4: {exc}") from exc

    @app.get("/invoice/m4-issued")
    def m4_issued(authorization: str | None = Header(None)):
        """Danh sach da xuat: chi don DA_XUAT_HD."""
        sb = _sb()
        actor = resolve_actor(sb, authorization)
        _require_ops(actor)

        try:
            res = (
                sb.table("don_hang")
                .select("*, khach_hang(*)")
                .eq("trang_thai_thu_tuc", TRANG_THAI_DA_XUAT_HD)
                .order("m3_approved_at", desc=True)
                .execute()
            )
            out = []
            for row in res.data or []:
                kh = row.pop("khach_hang", None) if isinstance(row, dict) else None
                out.append(_row_to_invoice_order(row, kh))
            return {"orders": out, "count": len(out)}
        except HTTPException:
            raise
        except Exception as exc:
            raise HTTPException(500, f"Loi lay danh sach da xuat M4: {exc}") from exc

    # ------------------------------------------------------------------
    # POST /invoice/m4-cancel
    # ------------------------------------------------------------------
    @app.post("/invoice/m4-cancel")
    def m4_cancel(body: M4CancelBody, authorization: str | None = Header(None)):
        """Revert đơn từ CHO_XUAT_HD về CHO_XAC_NHAN (hủy queue)."""
        sb = _sb()
        actor = resolve_actor(sb, authorization)
        _require_ops(actor)
        try:
            res = (
                sb.table("don_hang")
                .update(
                    {
                        "trang_thai_thu_tuc": TRANG_THAI_CHO_XAC_NHAN,
                        "m3_approved_at": None,
                        "tax_product_name": None,
                        "crm_order_id": None,
                    }
                )
                .eq("id", body.id)
                .eq("trang_thai_thu_tuc", TRANG_THAI_CHO_XUAT_HD)
                .execute()
            )
            if not res.data:
                raise HTTPException(
                    404, "Đơn không tìm thấy hoặc không ở trạng thái CHO_XUAT_HD"
                )
            return {"ok": True, "id": body.id}
        except HTTPException:
            raise
        except Exception as exc:
            raise HTTPException(500, f"Lỗi hủy queue M4: {exc}") from exc

    # ------------------------------------------------------------------
    # POST /invoice/export-batch  ← quan trọng nhất
    # ------------------------------------------------------------------
    @app.post("/invoice/export-batch")
    def export_batch(body: ExportBatchBody, authorization: str | None = Header(None)):
        """Xuất hóa đơn thuế.

        Mô tả luồng:
          - Lấy đơn hàng
          - Cấp sequences
          - Gán mã trong bộ nhớ
          - Tạo file Excel và nén ZIP
          - Cập nhật trạng thái trong cơ sở dữ liệu
          - Ghi lịch sử batch
          - Trả kết quả
        """
        sb = _sb()
        actor = resolve_actor(sb, authorization)
        _require_ops(actor)
        requested_ids = [str(x).strip() for x in (body.order_ids or []) if str(x).strip()]
        requested_ids = list(dict.fromkeys(requested_ids))

        # 1. Lấy toàn bộ đơn CHO_XUAT_HD
        try:
            res = (
                sb.table("don_hang")
                .select("*, khach_hang(*)")
                .eq("trang_thai_thu_tuc", TRANG_THAI_CHO_XUAT_HD)
                .order("m3_approved_at", desc=False)
                .execute()
            )
        except Exception as exc:
            raise HTTPException(500, f"Lỗi đọc queue xuất hóa đơn: {exc}") from exc

        raw_orders = [row for row in (res.data or []) if not row.get("tax_invoice_code")]
        if requested_ids:
            raw_orders = [row for row in raw_orders if str(row.get("id") or "") in requested_ids]
            found_ids = {str(row.get("id")) for row in raw_orders if row.get("id")}
            missing_ids = [oid for oid in requested_ids if oid not in found_ids]
            if missing_ids:
                raise HTTPException(
                    409,
                    "Mot so don khong con o trang thai cho xuat hoac da duoc xuat: "
                    + ", ".join(missing_ids),
                )
        if not raw_orders:
            raise HTTPException(400, "Hàng đợi trống — không có đơn nào để xuất hóa đơn")

        n = len(raw_orders)
        today = date.today()
        date_key = today.strftime("%d%m%y")  # DDMMYY — reset daily

        # 2. Cấp sequences (batch)
        try:
            inv_start, prod_start = _alloc_sequences(sb, n, date_key)
        except HTTPException:
            raise
        except Exception as exc:
            raise HTTPException(500, f"Lỗi cấp sequence hóa đơn: {exc}") from exc

        # 3. Gán mã IN-MEMORY (chưa update DB — OTHER-05)
        enriched: list[dict[str, Any]] = []
        pending_updates: list[dict[str, Any]] = []

        for i, row in enumerate(raw_orders):
            kh = row.pop("khach_hang", None) if isinstance(row, dict) else None
            tax_invoice_code = f"M{date_key}{inv_start + i + 1:03d}"
            tax_product_code = f"PF{prod_start + i + 1:06d}"

            order_data = _row_to_invoice_order(row, kh)
            order_data["taxInvoiceCode"] = tax_invoice_code
            order_data["taxProductCode"] = tax_product_code
            enriched.append(order_data)
            pending_updates.append({
                "id": row["id"],
                "tax_invoice_code": tax_invoice_code,
                "tax_product_code": tax_product_code,
            })

        # 4. Tạo 3 file Excel trong bộ nhớ
        try:
            excel_orders = _build_excel_orders(enriched)
            excel_customers = _build_excel_customers(enriched)
            excel_products = _build_excel_products(enriched)
        except HTTPException:
            raise
        except Exception as exc:
            raise HTTPException(500, f"Lỗi tạo file Excel: {exc}") from exc

        # 5. Nén ZIP ra file tạm
        batch_label = today.strftime("%Y%m%d")
        tmp = tempfile.NamedTemporaryFile(prefix="palfish_invoice_export_", suffix=".zip", delete=False)
        tmp_path = tmp.name
        tmp.close()
        try:
            with zipfile.ZipFile(tmp_path, "w", zipfile.ZIP_DEFLATED) as zf:
                zf.writestr(f"01_1don_hang_{batch_label}.xlsx", excel_orders)
                zf.writestr(f"02_khach_hang1_{batch_label}.xlsx", excel_customers)
                zf.writestr(f"03_sanpham1_{batch_label}.xlsx", excel_products)
        except Exception:
            os.unlink(tmp_path)
            raise

        # 6. ZIP thành công → BULK UPDATE DB (Atomic Transaction)
        try:
            bulk_data = [
                {
                    "id": pu["id"],
                    "tax_invoice_code": pu["tax_invoice_code"],
                    "tax_product_code": pu["tax_product_code"],
                    "trang_thai_thu_tuc": TRANG_THAI_DA_XUAT_HD,
                }
                for pu in pending_updates
            ]
            upd_res = sb.table("don_hang").upsert(bulk_data).execute()
            if not upd_res.data:
                raise HTTPException(500, "Không lưu được trạng thái xuất hóa đơn.")
        except HTTPException:
            os.unlink(tmp_path)
            raise
        except Exception as exc:
            os.unlink(tmp_path)
            raise HTTPException(500, f"Lỗi cập nhật CSDL: {exc}") from exc

        order_ids = [pu["id"] for pu in pending_updates]

        # 7. Ghi lịch sử batch (non-fatal)
        try:
            sb.table("export_batches").insert(
                {
                    "batch_date": today.isoformat(),
                    "order_ids": order_ids,
                    "order_count": n,
                    "created_by": actor.email,
                }
            ).execute()
        except Exception as exc:
            print(f"[invoice] Ghi export_batches thất bại (non-fatal): {exc}")

        # 8. Trả ZIP
        filename = f"hoa_don_thue_{batch_label}_{n}don.zip"
        return FileResponse(
            tmp_path,
            media_type="application/zip",
            headers={
                "Content-Disposition": f'attachment; filename="{filename}"',
                "Access-Control-Expose-Headers": "Content-Disposition, X-Batch-Count, X-Batch-Date",
                "X-Batch-Count": str(n),
                "X-Batch-Date": batch_label,
            },
            background=BackgroundTask(os.unlink, tmp_path),
        )
