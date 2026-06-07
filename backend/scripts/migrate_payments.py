#!/usr/bin/env python3
"""One-off migration: Google Sheet export (.xlsx) → payments module tables.

Sheets: SM Hanoi, HCM REV, Danang REV (Danang column map — verify before --apply).

Default: --dry-run. Use --apply to write DB.
"""

from __future__ import annotations

import argparse
import csv
import os
import sys
import time
from dataclasses import dataclass, field
from typing import Callable
from datetime import datetime, timezone
from decimal import Decimal
from pathlib import Path
from typing import Any

ROOT = Path(__file__).resolve().parents[2]
BACKEND_DIR = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(BACKEND_DIR))

if hasattr(sys.stdout, "reconfigure"):
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass

from dotenv import load_dotenv
from payment_logic import compute_gmv_final
from sheet_row_parsers import (
    parse_danang_row,
    parse_hcm_rev_row,
    parse_sm_hanoi_row,
)

SHEET_TABS = ("SM Hanoi", "HCM REV", "Danang REV")
DEFAULT_XLSX = Path(r"c:\Users\ductd\Downloads\All File Thu Hiền.xlsx")
INSERT_BATCH = 50
MASTER_BATCH = 100
SUPABASE_MAX_ATTEMPTS = 5

TYPE_TO_CHANNEL: dict[str, str] = {
    "refer": "Referral",
    "resell": "Renewal",
    "renewal": "Renewal",
    "广告": "Ads",
    "gd": "Public",
    "livestream": "Lives",
    "fb": "Ads",
    "koc": "KOC",
    "offline": "Offline",
}


def _log(msg: str) -> None:
    print(msg, flush=True)


def _short_code(name: str) -> str:
    parts = [p for p in name.replace(".", " ").split() if p]
    if not parts:
        return ""
    if len(parts) == 1:
        return parts[0][:4].title()
    return "".join(p[0].upper() for p in parts if p[0].isalpha())[:6]


def _channel_type(raw: str | None) -> str:
    if not raw:
        return "Other"
    key = raw.strip().lower()
    return TYPE_TO_CHANNEL.get(key, raw.strip())


def _dict_to_payment(raw: dict[str, Any], *, tab: str) -> ParsedPayment | None:
    pay_time = raw["pay_time"]
    if pay_time.tzinfo is None:
        pay_time = pay_time.replace(tzinfo=timezone.utc)
    vnd = int(raw["real_pay_vnd"])
    gmv_rmb = raw.get("gmv_rmb")
    gmv_final = float(
        compute_gmv_final(
            pay_time,
            Decimal(vnd),
            Decimal(str(gmv_rmb)) if gmv_rmb is not None else None,
        )
    )
    return ParsedPayment(
        uid=raw["uid"],
        pay_time=pay_time,
        real_pay_vnd=vnd,
        gmv_rmb=float(gmv_rmb) if gmv_rmb is not None else None,
        gmv_final=gmv_final,
        team=str(raw["team"]),
        sale_name=str(raw["sale_name"]),
        package_name=raw.get("package_name"),
        fixed=raw.get("fixed"),
        channel_type_raw=raw.get("channel_type_raw"),
        channel_code=raw.get("channel_code"),
        payment_seq=raw.get("payment_seq"),
        note=raw.get("note"),
        bank_day=raw.get("bank_day"),
        customer_name=raw.get("customer_name"),
        customer_phone=raw.get("customer_phone"),
        source_tab=tab,
    )


@dataclass
class ParsedPayment:
    uid: str
    pay_time: datetime
    real_pay_vnd: int
    gmv_rmb: float | None
    gmv_final: float
    team: str
    sale_name: str
    package_name: str | None
    fixed: str | None
    channel_type_raw: str | None
    channel_code: str | None
    payment_seq: str | None
    note: str | None
    bank_day: str | None
    customer_name: str | None
    customer_phone: str | None
    source_tab: str


@dataclass
class MigrationState:
    customers: dict[str, dict[str, Any]] = field(default_factory=dict)
    sales: dict[str, dict[str, Any]] = field(default_factory=dict)
    channels: dict[tuple[str, str], dict[str, Any]] = field(default_factory=dict)
    packages: dict[str, dict[str, Any]] = field(default_factory=dict)
    payments: list[ParsedPayment] = field(default_factory=list)
    duplicates: list[str] = field(default_factory=list)
    skipped: int = 0


def read_xlsx_rows(path: Path, tab: str) -> list[list[Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise SystemExit("Cần openpyxl: pip install openpyxl") from exc
    wb = load_workbook(path, read_only=True, data_only=True)
    if tab not in wb.sheetnames:
        _log(f"  ⚠ Sheet «{tab}» không có trong {path.name} — bỏ qua")
        return []
    ws = wb[tab]
    rows: list[list[Any]] = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        rows.append(list(row))
    wb.close()
    return rows


def collect_from_xlsx(path: Path, *, limit: int = 0) -> MigrationState:
    state = MigrationState()
    seen_keys: set[tuple[str, str, int]] = set()

    for tab in SHEET_TABS:
        rows = read_xlsx_rows(path, tab)
        if not rows:
            continue
        _log(f"  «{tab}»: {len(rows)} dòng")
        count = 0
        for row in rows:
            if tab == "Danang REV":
                raw = parse_danang_row(row)
                parsed = _dict_to_payment(raw, tab=tab) if raw else None
            elif tab == "HCM REV":
                raw = parse_hcm_rev_row(row)
                parsed = _dict_to_payment(raw, tab=tab) if raw else None
            else:
                raw = parse_sm_hanoi_row(row)
                parsed = _dict_to_payment(raw, tab=tab) if raw else None
            if not parsed:
                state.skipped += 1
                continue
            key = (
                parsed.uid,
                parsed.pay_time.isoformat(),
                parsed.real_pay_vnd,
            )
            if key in seen_keys:
                state.duplicates.append(
                    f"{tab}: uid={parsed.uid} pay_time={parsed.pay_time} vnd={parsed.real_pay_vnd}"
                )
                continue
            seen_keys.add(key)
            state.payments.append(parsed)
            _accumulate_masters(state, parsed)
            count += 1
            if limit and count >= limit:
                break
        if limit and len(state.payments) >= limit:
            break
    return state


def _accumulate_masters(state: MigrationState, p: ParsedPayment) -> None:
    if p.uid not in state.customers:
        state.customers[p.uid] = {
            "uid": p.uid,
            "full_name": p.customer_name,
            "phone": p.customer_phone,
            "first_seen": (p.bank_day or p.pay_time.date().isoformat()),
        }
    sale_key = p.sale_name.strip().lower()
    if sale_key and sale_key not in state.sales:
        state.sales[sale_key] = {
            "full_name": p.sale_name,
            "short_code": _short_code(p.sale_name),
            "team": p.team,
            "khoi": None,
            "active": True,
        }
    ch_key = (_channel_type(p.channel_type_raw), (p.channel_code or "").strip())
    if ch_key not in state.channels:
        state.channels[ch_key] = {
            "channel_code": p.channel_code,
            "name": p.channel_type_raw or ch_key[0],
            "type": ch_key[0],
        }
    if p.package_name and p.package_name not in state.packages:
        state.packages[p.package_name] = {
            "name": p.package_name,
            "fixed": p.fixed,
        }


def write_sales_review_csv(state: MigrationState, out_path: Path) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with out_path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(
            f,
            fieldnames=["full_name", "short_code", "team", "khoi", "active"],
        )
        w.writeheader()
        for row in sorted(state.sales.values(), key=lambda r: r["full_name"]):
            w.writerow(row)
    _log(f"  Sales review CSV: {out_path}")


def _is_retryable(exc: BaseException) -> bool:
    name = type(exc).__name__
    if name in (
        "RemoteProtocolError",
        "ReadTimeout",
        "ConnectTimeout",
        "ConnectError",
        "NetworkError",
        "WriteError",
        "PoolTimeout",
    ):
        return True
    msg = str(exc)
    return any(
        token in msg
        for token in (
            "ConnectionTerminated",
            "connection reset",
            "Server disconnected",
            "503",
            "502",
            "504",
        )
    )


def _sb_exec(fn: Callable[[], Any], *, label: str = "Supabase") -> Any:
    last: BaseException | None = None
    for attempt in range(1, SUPABASE_MAX_ATTEMPTS + 1):
        try:
            return fn()
        except Exception as exc:
            last = exc
            if not _is_retryable(exc) or attempt >= SUPABASE_MAX_ATTEMPTS:
                raise
            wait = min(2**attempt, 30)
            _log(f"  {label} retry {attempt}/{SUPABASE_MAX_ATTEMPTS} sau {wait}s: {exc}")
            time.sleep(wait)
    assert last is not None
    raise last


def _chunked(items: list[Any], size: int):
    for i in range(0, len(items), size):
        yield items[i : i + size]


def print_summary(state: MigrationState) -> None:
    by_tab: dict[str, int] = {}
    by_team: dict[str, int] = {}
    for p in state.payments:
        by_tab[p.source_tab] = by_tab.get(p.source_tab, 0) + 1
        by_team[p.team] = by_team.get(p.team, 0) + 1

    _log("")
    _log("=== Migration summary (dry-run / plan) ===")
    _log(f"  customers: {len(state.customers)}")
    _log(f"  sales:     {len(state.sales)}")
    _log(f"  channels:  {len(state.channels)}")
    _log(f"  packages:  {len(state.packages)}")
    _log(f"  payments:  {len(state.payments)}")
    _log(f"  skipped:   {state.skipped}")
    _log(f"  duplicates (skipped): {len(state.duplicates)}")
    _log("  payments by sheet:")
    for tab, n in sorted(by_tab.items()):
        _log(f"    {tab}: {n}")
    _log("  payments by team (top 10):")
    for team, n in sorted(by_team.items(), key=lambda x: -x[1])[:10]:
        _log(f"    {team}: {n}")
    if state.duplicates[:5]:
        for d in state.duplicates[:5]:
            _log(f"    dup: {d}")


def _load_existing_sales(sb) -> dict[str, str]:
    out: dict[str, str] = {}
    offset = 0
    while True:
        res = _sb_exec(
            lambda off=offset: (
                sb.table("sales").select("id, full_name").range(off, off + 999).execute()
            ),
            label="load sales",
        )
        chunk = res.data or []
        if not chunk:
            break
        for r in chunk:
            out[str(r["full_name"]).strip().lower()] = r["id"]
        if len(chunk) < 1000:
            break
        offset += 1000
    return out


def _load_existing_channels(sb) -> dict[tuple[str, str], str]:
    out: dict[tuple[str, str], str] = {}
    offset = 0
    while True:
        res = _sb_exec(
            lambda off=offset: (
                sb.table("channels")
                .select("id, type, channel_code")
                .range(off, off + 999)
                .execute()
            ),
            label="load channels",
        )
        chunk = res.data or []
        if not chunk:
            break
        for r in chunk:
            out[(r.get("type") or "Other", (r.get("channel_code") or "").strip())] = r["id"]
        if len(chunk) < 1000:
            break
        offset += 1000
    return out


def _load_existing_packages(sb) -> dict[str, str]:
    out: dict[str, str] = {}
    offset = 0
    while True:
        res = _sb_exec(
            lambda off=offset: (
                sb.table("packages").select("id, name").range(off, off + 999).execute()
            ),
            label="load packages",
        )
        chunk = res.data or []
        if not chunk:
            break
        for r in chunk:
            out[r["name"]] = r["id"]
        if len(chunk) < 1000:
            break
        offset += 1000
    return out


def apply_to_supabase(state: MigrationState, sb) -> None:
    sale_id_by_name = _load_existing_sales(sb)
    channel_id_by_key = _load_existing_channels(sb)
    package_id_by_name = _load_existing_packages(sb)

    cust_rows = list(state.customers.values())
    _log(f"  customers upsert: {len(cust_rows)}")
    for chunk in _chunked(cust_rows, MASTER_BATCH):
        _sb_exec(
            lambda c=chunk: sb.table("customers").upsert(c, on_conflict="uid").execute(),
            label="customers upsert",
        )

    sales_to_insert = [
        row for key, row in state.sales.items() if key not in sale_id_by_name
    ]
    _log(f"  sales insert: {len(sales_to_insert)} (existing {len(sale_id_by_name)})")
    for chunk in _chunked(sales_to_insert, MASTER_BATCH):
        res = _sb_exec(
            lambda c=chunk: sb.table("sales").insert(c).execute(),
            label="sales insert",
        )
        for r in res.data or []:
            sale_id_by_name[str(r["full_name"]).strip().lower()] = r["id"]
    sale_id_by_name = _load_existing_sales(sb)

    channels_to_insert = [
        row for key, row in state.channels.items() if key not in channel_id_by_key
    ]
    _log(f"  channels insert: {len(channels_to_insert)}")
    for chunk in _chunked(channels_to_insert, MASTER_BATCH):
        res = _sb_exec(
            lambda c=chunk: sb.table("channels").insert(c).execute(),
            label="channels insert",
        )
        for r in res.data or []:
            channel_id_by_key[
                (r.get("type") or "Other", (r.get("channel_code") or "").strip())
            ] = r["id"]
    channel_id_by_key = _load_existing_channels(sb)

    packages_to_insert = [
        row for name, row in state.packages.items() if name not in package_id_by_name
    ]
    _log(f"  packages insert: {len(packages_to_insert)}")
    for chunk in _chunked(packages_to_insert, MASTER_BATCH):
        res = _sb_exec(
            lambda c=chunk: sb.table("packages").insert(c).execute(),
            label="packages insert",
        )
        for r in res.data or []:
            package_id_by_name[r["name"]] = r["id"]
    package_id_by_name = _load_existing_packages(sb)

    batch: list[dict[str, Any]] = []
    inserted = 0
    errors = 0

    def flush_batch() -> None:
        nonlocal inserted, errors
        if not batch:
            return
        try:
            _sb_exec(
                lambda b=batch: sb.table("payments").insert(b).execute(),
                label="payments batch",
            )
            inserted += len(batch)
        except Exception as exc:
            if "payments_bizkey" in str(exc).lower() or "duplicate" in str(exc).lower():
                for row in batch:
                    try:
                        _sb_exec(
                            lambda r=row: sb.table("payments").insert(r).execute(),
                            label="payments row",
                        )
                        inserted += 1
                    except Exception as row_exc:
                        if "duplicate" in str(row_exc).lower() or "payments_bizkey" in str(row_exc).lower():
                            state.duplicates.append(f"DB dup uid={row.get('uid')}")
                        else:
                            errors += 1
                            _log(f"  ⚠ insert lỗi uid={row.get('uid')}: {row_exc}")
            else:
                errors += len(batch)
                _log(f"  ⚠ batch insert lỗi ({len(batch)} rows): {exc}")
        batch.clear()

    for p in state.payments:
        sale_id = sale_id_by_name.get(p.sale_name.strip().lower())
        if not sale_id:
            _log(f"  ⚠ Bỏ qua payment — không có sale_id cho «{p.sale_name}»")
            errors += 1
            continue
        ch_key = (_channel_type(p.channel_type_raw), (p.channel_code or "").strip())
        channel_id = channel_id_by_key.get(ch_key)
        package_id = package_id_by_name.get(p.package_name or "")
        batch.append(
            {
                "uid": p.uid,
                "pay_time": p.pay_time.isoformat(),
                "bank_day": p.bank_day,
                "package_id": package_id,
                "payment_seq": p.payment_seq,
                "real_pay_vnd": p.real_pay_vnd,
                "gmv_rmb": p.gmv_rmb,
                "gmv_final": p.gmv_final,
                "channel_id": channel_id,
                "sale_id": sale_id,
                "team": p.team,
                "status": "active",
                "note": p.note,
                "bank_matched": False,
                "crm_activated": False,
                "crm_order_id": None,
            }
        )
        if len(batch) >= INSERT_BATCH:
            flush_batch()
            if inserted and inserted % 1000 == 0:
                _log(f"  … {inserted} payments inserted")
    flush_batch()
    _log(f"  Inserted payments: {inserted}, errors: {errors}")


def main() -> None:
    parser = argparse.ArgumentParser(description="Migrate Google Sheet export → payments tables")
    parser.add_argument(
        "--input",
        type=Path,
        default=DEFAULT_XLSX,
        help="Path to .xlsx with sheets SM Hanoi, HCM REV, Danang REV",
    )
    parser.add_argument(
        "--apply",
        action="store_true",
        help="Write to Supabase (default: dry-run only)",
    )
    parser.add_argument("--limit", type=int, default=0, help="Max payments per run (0 = all)")
    parser.add_argument(
        "--review-sales-csv",
        type=Path,
        default=BACKEND_DIR / "scripts" / "out" / "payments_sales_review.csv",
        help="Output CSV for manual sales master review",
    )
    args = parser.parse_args()

    load_dotenv(BACKEND_DIR / ".env")
    load_dotenv(BACKEND_DIR / ".env.sandbox")
    load_dotenv(ROOT / ".env")
    load_dotenv(ROOT / ".env.sandbox")

    if not args.input or not args.input.is_file():
        _log("Chưa có file .xlsx — chạy template/dry-run mode.")
        _log("  Export Google Sheet → .xlsx (3 tabs) rồi chạy:")
        _log("  python backend/scripts/migrate_payments.py --input path/to/export.xlsx")
        _log("")
        _log("Column mappers: sheet_row_parsers (SM Hanoi, HCM REV, Danang REV — verified)")
        _log(f"  Default input: {DEFAULT_XLSX}")
        sys.exit(0 if not args.apply else 1)

    _log(f"Đọc {args.input} …")

    sb = None
    if os.getenv("SUPABASE_URL") and os.getenv("SUPABASE_SERVICE_ROLE_KEY"):
        from supabase import create_client

        sb = create_client(
            os.environ["SUPABASE_URL"].strip(),
            os.environ["SUPABASE_SERVICE_ROLE_KEY"].strip(),
        )

    state = collect_from_xlsx(args.input, limit=args.limit)
    write_sales_review_csv(state, args.review_sales_csv)
    print_summary(state)

    if args.apply:
        if not sb:
            _log("[ERROR] --apply cần SUPABASE_URL + SUPABASE_SERVICE_ROLE_KEY trong .env")
            sys.exit(1)
        _log("")
        _log("=== APPLY to Supabase ===")
        apply_to_supabase(state, sb)
    else:
        _log("")
        _log("Dry-run xong. Review sales CSV, rồi chạy lại với --apply.")


if __name__ == "__main__":
    main()
