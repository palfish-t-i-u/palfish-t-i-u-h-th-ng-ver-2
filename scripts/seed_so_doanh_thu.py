#!/usr/bin/env python3
"""Seed / backfill bảng so_doanh_thu.

Cần SUPABASE_URL + SUPABASE_SERVICE_ROLE_KEY (backend/.env hoặc env).

Ví dụ:
  # Backfill đơn M3 đã approve trước khi có Module 5
  python scripts/seed_so_doanh_thu.py --backfill-m3

  # Import lịch sử Excel (sheet HN + HCM), thử 500 dòng
  python scripts/seed_so_doanh_thu.py --xlsx "e:\\PalFish\\DA\\HNxHCM GMV.xlsx" --limit 500

  # Xem trước, không ghi DB
  python scripts/seed_so_doanh_thu.py --xlsx "..." --limit 10 --dry-run
"""

from __future__ import annotations

import argparse
import os
import sys
from datetime import date, datetime
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "backend"))

DEFAULT_TY_GIA = Decimal("3700")

TEAM_FROM_EXCEL: dict[str, str] = {
    "In-house": "Inhouse 1",
    "In-house 2": "Inhouse 2",
    "HCM team": "HCM (Online)",
    "Linh Dam": "Linh Dam (Store)",
    "Offline": "Offline",
    "An Binh": "An Binh (Store)",
}

PIVOT_FROM_APP: dict[str, str] = {
    "Inhouse 1": "HN inhouse",
    "Inhouse 2": "HN inhouse 2",
    "HCM (Online)": "HCM team",
    "Linh Dam (Store)": "Linh Dam",
    "Offline": "Offline",
    "An Binh (Store)": "An Binh",
}


def _load_dotenv() -> None:
    env_path = ROOT / "backend" / ".env"
    if env_path.exists():
        try:
            from dotenv import load_dotenv

            load_dotenv(env_path)
        except ImportError:
            pass


def _parse_date(value) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    s = str(value).strip()
    if not s:
        return None
    try:
        return date.fromisoformat(s[:10])
    except ValueError:
        return None


def _vnd_to_rmb(vnd: int, gmv_excel: float | None = None) -> float:
    if gmv_excel and float(gmv_excel) > 0:
        return round(float(gmv_excel), 2)
    if not vnd:
        return 0.0
    r = (Decimal(str(vnd)) / DEFAULT_TY_GIA).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    return float(r)


def _map_team(excel_team: str | None) -> tuple[str | None, str | None]:
    raw = (excel_team or "").strip()
    if not raw:
        return None, None
    app_team = TEAM_FROM_EXCEL.get(raw, raw)
    pivot = PIVOT_FROM_APP.get(app_team, raw)
    return app_team, pivot


def _row_from_excel(row: tuple, sheet: str) -> dict | None:
    if not row or len(row) < 10:
        return None
    vnd = int(float(row[9] or 0))
    if vnd <= 0:
        return None

    ngay = _parse_date(row[0]) or _parse_date(row[8])
    if not ngay:
        return None

    app_team, pivot = _map_team(str(row[17]).strip() if len(row) > 17 and row[17] else None)
    gmv = float(row[10]) if len(row) > 10 and row[10] else None

    uid_val = row[5]
    uid = str(int(uid_val)) if isinstance(uid_val, (int, float)) and uid_val else str(uid_val or "")

    return {
        "ngay_tien_ve": ngay.isoformat(),
        "bank_time": row[1].isoformat() if hasattr(row[1], "isoformat") else None,
        "gateway": str(row[2] or "") or None,
        "ten_khach": str(row[3] or "").strip(),
        "sdt": str(row[4] or "").strip() or None,
        "uid": uid or None,
        "goi_hoc": str(row[6] or "").strip() or None,
        "fixed_non_fixed": str(row[7] or "").strip() or None,
        "pay_time": row[8].isoformat() if isinstance(row[8], datetime) else None,
        "so_tien_vnd": vnd,
        "gmv_rmb": _vnd_to_rmb(vnd, gmv),
        "ty_gia_vnd_rmb": float(DEFAULT_TY_GIA),
        "payment_method": str(row[11] or "").strip() or None if len(row) > 11 else None,
        "loai": str(row[12] or "").strip() or None if len(row) > 12 else None,
        "loai_2": str(row[13] or "").strip() or None if len(row) > 13 else None,
        "sale_crm_name": str(row[14] or "").strip() or None if len(row) > 14 else None,
        "note": str(row[15] or "").strip() or None if len(row) > 15 else None,
        "note2": str(row[18] or "").strip() or None if len(row) > 18 else None,
        "team": app_team,
        "team_pivot_label": pivot,
        "loai_nhap": "tay",
        "created_by_email": f"import:{sheet}",
        "updated_by_email": f"import:{sheet}",
    }


def import_xlsx(sb, path: Path, limit: int, dry_run: bool) -> int:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    payloads: list[dict] = []
    for sheet in ("HN", "HCM"):
        if sheet not in wb.sheetnames:
            print(f"  skip sheet {sheet} (missing)")
            continue
        ws = wb[sheet]
        for row in ws.iter_rows(min_row=2, values_only=True):
            p = _row_from_excel(row, sheet)
            if p:
                payloads.append(p)
            if limit and len(payloads) >= limit:
                break
        if limit and len(payloads) >= limit:
            break

    print(f"Xlsx rows to insert: {len(payloads)} (limit={limit or 'none'})")
    if dry_run:
        for p in payloads[:5]:
            print("  sample:", p["ngay_tien_ve"], p["ten_khach"], p["so_tien_vnd"], p["team"])
        return 0

    inserted = 0
    batch = 100
    for i in range(0, len(payloads), batch):
        chunk = payloads[i : i + batch]
        sb.table("so_doanh_thu").insert(chunk).execute()
        inserted += len(chunk)
        print(f"  inserted {inserted}/{len(payloads)}")
    return inserted


def backfill_m3(sb) -> int:
    from revenue_routes import sync_ledger_from_m3_order

    res = (
        sb.table("don_hang")
        .select("id, ma_don_hang, trang_thai_thu_tuc, m3_approved_at")
        .in_("trang_thai_thu_tuc", ["CHO_XUAT_HD", "DA_XUAT_HD"])
        .execute()
    )
    orders = res.data or []
    print(f"M3 orders (CHO_XUAT_HD / DA_XUAT_HD): {len(orders)}")

    created = 0
    skipped = 0
    failed = 0
    for o in orders:
        oid = o["id"]
        rid = sync_ledger_from_m3_order(sb, oid, "backfill@m3")
        if rid:
            created += 1
            print(f"  + {o.get('ma_don_hang') or oid[:8]} → {rid[:8]}")
        else:
            existing = (
                sb.table("so_doanh_thu")
                .select("id")
                .eq("don_hang_id", oid)
                .eq("loai_nhap", "tu_dong")
                .limit(1)
                .execute()
            )
            if existing.data:
                skipped += 1
            else:
                failed += 1
                print(f"  ! failed {o.get('ma_don_hang') or oid[:8]}")

    print(f"Backfill done: created={created}, already_had={skipped}, failed={failed}")
    return created


def main() -> int:
    _load_dotenv()
    parser = argparse.ArgumentParser(description="Seed so_doanh_thu from Excel or M3 backfill")
    parser.add_argument("--xlsx", type=Path, help="Path to HNxHCM GMV.xlsx")
    parser.add_argument("--backfill-m3", action="store_true", help="Sync đơn M3 đã approve vào Sổ")
    parser.add_argument("--limit", type=int, default=0, help="Max rows import xlsx (0 = all)")
    parser.add_argument("--dry-run", action="store_true", help="Chỉ in sample, không insert")
    args = parser.parse_args()

    if not args.xlsx and not args.backfill_m3:
        parser.error("Cần --xlsx hoặc --backfill-m3")

    url = os.getenv("SUPABASE_URL", "").strip()
    key = os.getenv("SUPABASE_SERVICE_ROLE_KEY", "").strip()
    if not url or not key:
        print("Set SUPABASE_URL and SUPABASE_SERVICE_ROLE_KEY", file=sys.stderr)
        return 1

    from supabase import create_client

    sb = create_client(url, key)

    if args.backfill_m3:
        backfill_m3(sb)

    if args.xlsx:
        if not args.xlsx.exists():
            print(f"File not found: {args.xlsx}", file=sys.stderr)
            return 1
        import_xlsx(sb, args.xlsx, args.limit, args.dry_run)

    return 0


if __name__ == "__main__":
    sys.exit(main())
