#!/usr/bin/env python3
"""Month-by-month BC01 Inhouse 1 audit: Sổ vs HNxHCM GMV vs sale-level diffs."""

from __future__ import annotations

import os
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "backend"))

from dotenv import load_dotenv

load_dotenv(ROOT / "backend" / ".env")

from revenue_routes import _month_key, _parse_date, team_to_canonical  # noqa: E402

TEAM_FILTER = "Inhouse 1"
HN_PATH = Path(r"e:\PalFish\DA\HNxHCM GMV.xlsx")
SHEET_PATH = Path(r"e:\PalFish\DA\All File Thu Hiền.xlsx")

# From user screenshots (Google Sheet live — HN Inhouse 1 row Tổng)
SHEET_LIVE_TOTALS: dict[str, float] = {
    "2025/8": 1_505_332,
    "2025/9": 1_284_598,
    "2025/10": 1_278_255,
    "2025/11": 1_462_392,
    "2025/12": 1_591_663,
    "2026/2": 1_223_593,
    "2026/3": 1_798_894,  # image OCR ambiguous; user may mean 1,738,894
    "2026/4": 1_815_778,
    "2026/5": 1_009_516,
}


def fetch_so_full() -> list[dict]:
    from supabase import create_client

    sb = create_client(os.environ["SUPABASE_URL"], os.environ["SUPABASE_SERVICE_ROLE_KEY"])
    rows: list[dict] = []
    offset = 0
    while True:
        q = (
            sb.table("so_doanh_thu")
            .select("pay_time, ngay_tien_ve, gmv_rmb, sale_crm_name, team, team_pivot_label")
            .order("pay_time", desc=True)
            .order("id", desc=True)
        )
        chunk = q.range(offset, offset + 999).execute().data or []
        if not chunk:
            break
        rows.extend(chunk)
        if len(chunk) < 1000:
            break
        offset += 1000
    return rows


def so_monthly(rows: list[dict], *, month_field: str) -> dict[str, dict[str, float]]:
    """month_field: 'pay_time' | 'ngay_tien_ve'"""
    by_month: dict[str, dict[str, float]] = defaultdict(lambda: defaultdict(float))
    for r in rows:
        if team_to_canonical(r.get("team"), r.get("team_pivot_label")) != TEAM_FILTER:
            continue
        d = _parse_date(r.get(month_field))
        if not d:
            continue
        mk = _month_key(d)
        sale = (r.get("sale_crm_name") or "(Chưa gán sale)").strip()
        by_month[mk][sale] += float(r.get("gmv_rmb") or 0)
    return by_month


def load_hnxhcm_monthly() -> dict[str, dict[str, float]]:
    import openpyxl

    wb = openpyxl.load_workbook(HN_PATH, read_only=True, data_only=True)
    ws = wb["HNxHCM"]
    hdr = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    cols = {str(h).lower(): i for i, h in enumerate(hdr) if h}
    sales_i = cols["sales"]
    team_i = cols["team"]
    pay_i = cols["pay time"]
    gmv_i = next(i for i, h in enumerate(hdr) if h and "gmv" in str(h).lower())

    by_month: dict[str, dict[str, float]] = defaultdict(lambda: defaultdict(float))
    for row in ws.iter_rows(min_row=2, values_only=True):
        row = list(row)
        team = str(row[team_i] or "").strip()
        if team not in ("In-house", "Inhouse 1") and "inhouse" not in team.lower():
            continue
        pt = row[pay_i]
        if not isinstance(pt, datetime):
            continue
        sale = str(row[sales_i] or "").strip() or "(Chưa gán sale)"
        mk = _month_key(pt.date())
        by_month[mk][sale] += float(row[gmv_i] or 0)
    wb.close()
    return by_month


def load_xlsx_pivot_monthly() -> dict[str, dict[str, float]]:
    import openpyxl

    if not SHEET_PATH.exists():
        return {}
    wb = openpyxl.load_workbook(SHEET_PATH, read_only=True, data_only=True)
    ws = wb["HN Inhouse 1"]
    hdr = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
    month_cols = {f"{v.year}/{v.month}": i for i, v in enumerate(hdr) if hasattr(v, "year")}
    by_month: dict[str, dict[str, float]] = defaultdict(dict)
    for row in ws.iter_rows(min_row=4, values_only=True):
        row = list(row)
        if len(row) < 3 or not row[1]:
            continue
        sale = str(row[1]).strip()
        if sale.startswith("总") or sale == "销售":
            continue
        for mk, ci in month_cols.items():
            if ci < len(row) and row[ci]:
                v = float(row[ci] or 0)
                if v:
                    by_month[mk][sale] = v
    wb.close()
    return by_month


def month_sort(mk: str) -> tuple[int, int]:
    y, m = mk.split("/")
    return int(y), int(m)


def diff_sales(
    a: dict[str, float], b: dict[str, float], *, top: int = 8
) -> tuple[list[tuple[str, float, float, float]], float, float]:
    keys = set(a) | set(b)
    diffs = [(k, a.get(k, 0), b.get(k, 0), a.get(k, 0) - b.get(k, 0)) for k in keys]
    diffs = [d for d in diffs if abs(d[3]) > 0.5]
    diffs.sort(key=lambda x: -abs(x[3]))
    return diffs[:top], sum(a.values()), sum(b.values())


def main() -> int:
    print("Loading Sổ doanh thu…")
    rows = fetch_so_full()
    so_pay = so_monthly(rows, month_field="pay_time")
    so_ntv = so_monthly(rows, month_field="ngay_tien_ve")
    print("Loading HNxHCM GMV…")
    hn = load_hnxhcm_monthly()
    xlsx = load_xlsx_pivot_monthly()

    months = sorted(set(so_ntv) | set(hn) | set(SHEET_LIVE_TOTALS), key=month_sort)
    months = [m for m in months if m >= "2025/1" and m <= "2026/5"]

    print(f"\n{'Tháng':<8} {'Sổ(pay)':>12} {'Sổ(ntv)':>12} {'HNxHCM':>12} {'Sheet live':>12} {'Δ ntv-live':>12}")
    print("-" * 72)
    problem_months: list[str] = []
    for mk in months:
        t_pay = round(sum(so_pay.get(mk, {}).values()))
        t_ntv = round(sum(so_ntv.get(mk, {}).values()))
        t_hn = round(sum(hn.get(mk, {}).values()))
        t_live = SHEET_LIVE_TOTALS.get(mk)
        live_s = f"{t_live:,.0f}" if t_live is not None else "—"
        delta = t_ntv - t_live if t_live is not None else None
        delta_s = f"{delta:+,.0f}" if delta is not None else "—"
        flag = ""
        if t_live is not None and abs(delta or 0) > 100:
            flag = " *"
            problem_months.append(mk)
        if t_ntv != t_hn:
            flag += " HN≠Sổ"
        print(f"{mk:<8} {t_pay:>12,} {t_ntv:>12,} {t_hn:>12,} {live_s:>12} {delta_s:>12}{flag}")

    print("\n=== Chi tiết tháng user hỏi (Sổ ngay_tien_ve vs Sheet live) ===\n")
    check = ["2025/5", "2025/6", "2025/7", "2025/8", "2025/9", "2025/10", "2025/11", "2025/12", "2026/2", "2026/3", "2026/5"]
    for mk in check:
        live = SHEET_LIVE_TOTALS.get(mk)
        if live is None and mk in xlsx:
            live = round(sum(xlsx[mk].values()))
        so_map = so_ntv.get(mk, {})
        live_map = xlsx.get(mk, {})
        if not live_map and live:
            print(f"--- {mk}: Sổ={sum(so_map.values()):,.0f} | Sheet live total={live:,.0f} (no per-sale xlsx) ---")
        else:
            print(f"--- {mk}: Sổ={sum(so_map.values()):,.0f} | Sheet={live or sum(live_map.values()):,.0f} ---")
        # diff vs live: reconstruct from so only vs known sheet sales from xlsx when available
        if live_map:
            top, ta, tb = diff_sales(so_map, live_map)
            for sale, va, vb, d in top:
                print(f"  {sale}: Sổ={va:,.0f} Sheet={vb:,.0f} Δ={d:+,.0f}")
            only_so = {k: v for k, v in so_map.items() if k not in live_map and v > 1}
            only_sheet = {k: v for k, v in live_map.items() if k not in so_map and v > 1}
            if only_so:
                print(f"  Chỉ Sổ ({sum(only_so.values()):,.0f}):", ", ".join(f"{k} ({v:,.0f})" for k, v in sorted(only_so.items(), key=lambda x: -x[1])[:6]))
            if only_sheet:
                print(f"  Chỉ Sheet ({sum(only_sheet.values()):,.0f}):", ", ".join(f"{k} ({v:,.0f})" for k, v in sorted(only_sheet.items(), key=lambda x: -x[1])[:6]))
        print()

    # pay_time vs ngay_tien_ve drift
    drift_months = [mk for mk in months if abs(sum(so_pay.get(mk, {}).values()) - sum(so_ntv.get(mk, {}).values())) > 1]
    if drift_months:
        print("Tháng pay_time ≠ ngay_tien_ve (Sổ):", ", ".join(drift_months))
    else:
        print("pay_time và ngay_tien_ve cho cùng tháng bucket: khớp toàn bộ Inhouse 1.")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
